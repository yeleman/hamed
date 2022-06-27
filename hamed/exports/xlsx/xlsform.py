#!/usr/bin/env python
# -*- coding: utf-8 -*-
# vim: ai ts=4 sts=4 et sw=4 nu

import logging
import io

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from hamed.exports.xlsx import (
    IncorrectExcelFile,
    letter_to_column,
    indexes_to_coordinate,
)

logger = logging.getLogger(__name__)


def gen_xlsform(template_path, form_title, form_id):
    try:
        wb = load_workbook(template_path)
    except InvalidFileException:
        raise IncorrectExcelFile("Not a proper XLSX Template.")

    ws = wb["settings"]

    # find columns for those settings (or create)
    form_id_cell = ws.cell(row=2, column=get_setting_column(ws, "form_id"))
    form_title_cell = ws.cell(row=2, column=get_setting_column(ws, "form_title"))
    # set new values
    form_id_cell.value = form_id
    form_title_cell.value = form_title

    xform = io.BytesIO()
    wb.save(xform)
    xform.seek(0)  # make sure it's readable before returning

    return xform


def get_setting_column(ws, key, create_if_missing=True):
    last_col = 0
    for col in ws.iter_cols(min_row=0, max_row=0):
        for cell in col:
            if cell.value == key:
                return letter_to_column(cell.column)
        last_col = cell.column
    column = letter_to_column(last_col) + 1
    coordinate = indexes_to_coordinate(0, column)
    if create_if_missing:
        cell = ws[coordinate] = key
    return column
